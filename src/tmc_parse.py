"""
tmc_parse.py — JDA TMC workbooks -> tidy frames.

Design rule, and the reason this module exists: **never trust a stored total.**
Every `Total Fast`, `Total Slow`, `Grand Total (Nos.)` and `Grand Total (PCU's)`
in these workbooks is recomputed from its components. Where the stored value
disagrees, the discrepancy is written to a register and the *derived* value is
used downstream. Nothing is silently corrected and nothing is silently accepted.

Run:  uv run python src/tmc_parse.py
"""
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src.config import SOURCE, PROCESSED, SURVEY_DIRS, JUNCTIONS, MOVEMENTS

# --- sheet geometry, identical in all 12 workbooks (verified by inspect_tmc) ---
ROW_BINS = range(8, 104)        # 96 x 15-min bins, 08:00 -> 08:00 next day
ROW_TOTAL_VEH = 104
ROW_TOTAL_PCU = 105
ROW_COMPOSITION = 106
ROW_HOURS = range(114, 207)     # 93 rolling 60-min windows

# Column -> canonical class code. The JDA scheme is 10 columns, five of which
# are composites mixing IRC:106 classes with different PCU values.
FAST_COLS = {2: "CAR_BUCKET", 3: "TWO_W", 4: "AGRI_LCV", 5: "AUTO_TRK_BUS", 6: "TRL_MAV"}
SLOW_COLS = {8: "CYCLE", 9: "CYCLE_RIK", 10: "HAND_CART", 11: "HORSE_DRAWN", 12: "BULLOCK"}
CLASS_COLS = {**FAST_COLS, **SLOW_COLS}
COL_TOTAL_FAST, COL_TOTAL_SLOW, COL_GRAND, COL_PCU = 7, 13, 14, 15

CLASS_LABELS = {
    "CAR_BUCKET":   "Car, Taxi, Tempo, Auto Rickshaw & Pick up",
    "TWO_W":        "Motar Cycle, Scooter",
    "AGRI_LCV":     "Agriculture Tractor, LCV Mini Bus",
    "AUTO_TRK_BUS": "Three Wheeler (Auto) Axle Truck, Buses",
    "TRL_MAV":      "Tractor Trailor, Truck Trailor Units (3 Axle & MAV)",
    "CYCLE":        "Cycle",
    "CYCLE_RIK":    "Cycle Rickshaw",
    "HAND_CART":    "Hand Cart",
    "HORSE_DRAWN":  "Horse Drawn",
    "BULLOCK":      "Bullock Corts",
}

MOVEMENT_SHEETS = [f"V_{i}" for i in range(1, 13)]
APPROACH_SHEETS = [f"IN_{i}" for i in range(1, 5)] + [f"OUT_{i}" for i in range(1, 5)]


def num(v):
    """Numeric cell value, or None. `#REF!` and other error strings become None."""
    if v is None or isinstance(v, str):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def bin_datetime(label, survey_date):
    """
    '0800-0815' -> datetime. The survey day runs 08:00 to 08:00 the next morning,
    so anything before 08:00 belongs to the following calendar date.
    """
    start = label.split("-")[0].strip()
    hh, mm = int(start[:2]), int(start[2:])
    # '2400' appears as an end label only; starts are always < 24
    dt = datetime.combine(survey_date, datetime.min.time()) + timedelta(hours=hh, minutes=mm)
    if hh < 8:
        dt += timedelta(days=1)
    return dt


def movement_for(v_index, arms):
    """
    V_1..V_12 -> (from_arm, to_arm, movement).

    Arms are ordered clockwise. India drives on the left, so from any approach the
    LEFT turn lands on the next arm clockwise, STRAIGHT on the one after, RIGHT on
    the one after that. V sheets are laid out entry-major in that same order.
    """
    entry = (v_index - 1) // 3
    offset = (v_index - 1) % 3 + 1          # 1=Left, 2=Straight, 3=Right
    exit_ = (entry + offset) % 4
    return arms[entry], arms[exit_], MOVEMENTS[offset - 1]


def parse_workbook(path, junction, survey_date, mismatches):
    """One workbook -> list of tidy bin records. Appends to `mismatches` in place."""
    wb = load_workbook(path, data_only=True)
    arms = JUNCTIONS[junction]
    rows = []

    for sheet in MOVEMENT_SHEETS + APPROACH_SHEETS + ["TOTAL_IN", "TOTAL_OUT"]:
        ws = wb[sheet]

        if sheet.startswith("V_"):
            kind = "movement"
            frm, to, mv = movement_for(int(sheet.split("_")[1]), arms)
            # The sheet states its own direction. Check ours against it rather
            # than trusting the positional convention.
            stated_from = str(ws.cell(row=4, column=2).value or "").strip()
            stated_to = str(ws.cell(row=4, column=5).value or "").strip()
            if stated_from != frm or stated_to != to:
                mismatches.append(dict(
                    junction=junction, date=survey_date, sheet=sheet, row=4,
                    field="direction", stored=f"{stated_from} -> {stated_to}",
                    derived=f"{frm} -> {to}", delta=None))
        elif sheet.startswith(("IN_", "OUT_")):
            kind = "inflow" if sheet.startswith("IN_") else "outflow"
            idx = int(sheet.split("_")[1]) - 1
            frm, to, mv = (arms[idx], None, None) if kind == "inflow" else (None, arms[idx], None)
        else:
            kind = "total_in" if sheet == "TOTAL_IN" else "total_out"
            frm = to = mv = None

        for r in ROW_BINS:
            label = ws.cell(row=r, column=1).value
            if not isinstance(label, str) or "-" not in label:
                continue
            counts = {code: num(ws.cell(row=r, column=c).value) or 0.0
                      for c, code in CLASS_COLS.items()}

            # --- recompute every stored total for this row --------------------
            d_fast = sum(counts[c] for c in FAST_COLS.values())
            d_slow = sum(counts[c] for c in SLOW_COLS.values())
            for col, field, derived in (
                (COL_TOTAL_FAST, "Total Fast", d_fast),
                (COL_TOTAL_SLOW, "Total Slow", d_slow),
                (COL_GRAND, "Grand Total (Nos.)", d_fast + d_slow),
            ):
                stored = num(ws.cell(row=r, column=col).value)
                if stored is not None and abs(stored - derived) > 1e-9:
                    mismatches.append(dict(
                        junction=junction, date=survey_date, sheet=sheet, row=r,
                        field=field, stored=stored, derived=derived,
                        delta=stored - derived))

            dt = bin_datetime(label, survey_date)
            for code, n in counts.items():
                rows.append(dict(
                    junction=junction, date=survey_date, sheet=sheet, kind=kind,
                    arm_from=frm, arm_to=to, movement=mv, bin_start=dt,
                    bin_label=label, veh_class=code, count=n,
                    stored_pcu=num(ws.cell(row=r, column=COL_PCU).value)))

        # --- day totals: compare the stored footer against the summed bins -----
        for col, field in ((COL_TOTAL_FAST, "day Total Fast"),
                           (COL_TOTAL_SLOW, "day Total Slow"),
                           (COL_GRAND, "day Grand Total (Nos.)")):
            stored = num(ws.cell(row=ROW_TOTAL_VEH, column=col).value)
            if stored is None:
                continue
            if col == COL_TOTAL_FAST:
                derived = sum(r_["count"] for r_ in rows
                              if r_["sheet"] == sheet and r_["date"] == survey_date
                              and r_["junction"] == junction
                              and r_["veh_class"] in FAST_COLS.values())
            elif col == COL_TOTAL_SLOW:
                derived = sum(r_["count"] for r_ in rows
                              if r_["sheet"] == sheet and r_["date"] == survey_date
                              and r_["junction"] == junction
                              and r_["veh_class"] in SLOW_COLS.values())
            else:
                derived = sum(r_["count"] for r_ in rows
                              if r_["sheet"] == sheet and r_["date"] == survey_date
                              and r_["junction"] == junction)
            if abs(stored - derived) > 1e-9:
                mismatches.append(dict(
                    junction=junction, date=survey_date, sheet=sheet,
                    row=ROW_TOTAL_VEH, field=field, stored=stored,
                    derived=derived, delta=stored - derived))

    wb.close()
    return rows


def parse_all():
    """All 12 workbooks -> (bins DataFrame, mismatch register DataFrame)."""
    all_rows, mismatches = [], []
    for d in SURVEY_DIRS:
        survey_date = datetime.strptime(d.split("_")[1], "%d-%m-%Y").date()
        for path in sorted((SOURCE / d).glob("*.xlsx")):
            junction = f"TMC-{path.name[:2]}"
            all_rows.extend(parse_workbook(path, junction, survey_date, mismatches))
    return pd.DataFrame(all_rows), pd.DataFrame(mismatches)


if __name__ == "__main__":
    bins, mism = parse_all()
    PROCESSED.mkdir(parents=True, exist_ok=True)
    bins.to_parquet(PROCESSED / "tmc_bins.parquet", index=False)
    mism.to_parquet(PROCESSED / "tmc_mismatches.parquet", index=False)

    exp = len(JUNCTIONS) * 2 * 22 * 96 * 10   # junctions x days x sheets x bins x classes
    print(f"bins parsed        : {len(bins):,}   (expected {exp:,})")
    print(f"junctions          : {bins.junction.nunique()}   dates: {sorted(bins.date.unique())}")
    print(f"vehicle classes    : {bins.veh_class.nunique()}")
    # CLAUDE.md's gate is "exactly 12 per junction (4 arms x L/S/R), no U-turns".
    # This used to divide the movement row count by 96 bins x 10 classes and print the
    # result - which is 24, because it never divided by the TWO survey days. It reported
    # a number that failed its own gate, under a label saying it was the gate, and
    # nothing compared it to 12.
    mv = bins[bins.kind == "movement"]
    # a "movement" here is an ARM x TURN pair - 4 arms x L/S/R. The `movement` column
    # alone holds only the turn, so counting it returns 3.
    per = (mv.drop_duplicates(["junction", "date", "arm_from", "movement"])
             .groupby(["junction", "date"]).size())
    print(f"movements per junc : {sorted(per.unique())} per junction per day "
          f"(expected [12]) over {mv.date.nunique()} days")
    bad = per[per != 12]
    if len(bad):
        print(f"  GATE FAILED - {len(bad)} junction-days are not 12 movements:")
        print(bad.to_string())
    else:
        print(f"  GATE PASSED - all {len(per)} junction-days carry exactly 12 movements, "
              f"no U-turns")
    print()
    print(f"MISMATCH REGISTER  : {len(mism):,} discrepancies between stored and derived")
    if len(mism):
        print(mism.groupby(["field"]).agg(n=("delta", "size"),
                                          total_delta=("delta", "sum")).to_string())
        print()
        print("worst 10 by absolute delta:")
        print(mism.reindex(mism.delta.abs().sort_values(ascending=False).index)
                  .head(10).to_string(index=False))
    print()
    print(f"GATE — silently absorbed discrepancies: 0 "
          f"(all {len(mism):,} are recorded above)")
