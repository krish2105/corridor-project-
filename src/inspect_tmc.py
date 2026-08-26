"""
inspect_tmc.py — Structure probe for the JDA classified TMC workbooks.

Inventory only. This deliberately does NOT reshape, clean, coerce or tidy anything:
the point is to see the sheet exactly as the consultant laid it out, merged headers
and all, before deciding how to parse it.

Run:  uv run python src/inspect_tmc.py
"""
import re
import sys
from collections import Counter, defaultdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(__import__("pathlib").Path(__file__).resolve().parent.parent))
from src.tmc_parse import workbooks
from src.config import SOURCE, SURVEY_DIRS

# The file the deep probe runs against.
PRIMARY = SOURCE / "INT_11-05-2026" / "01_TMC (11-05-2026).xlsx"

RAW_ROWS = 30          # how many rows of the first sheet to dump verbatim
BAR = "=" * 78
SUB = "-" * 78

# --- discovery vocabularies -------------------------------------------------
# Keyword matching, not a fixed schema. We are finding out what labels exist,
# so anything that looks like a class/movement/time gets surfaced and the
# classification is reported as a guess, not applied as a transform.
VEHICLE_HINTS = (
    "car", "taxi", "tempo", "rickshaw", "pick up", "pickup", "motar", "motor",
    "scooter", "cycle", "tractor", "trailor", "trailer", "lcv", "mav", "bus",
    "truck", "axle", "hand cart", "horse", "bullock", "wheeler", "goods",
    "vehicle", "fast", "slow", "others", "mini", "e-rickshaw", "tw",
)
MOVEMENT_HINTS = (
    "left", "right", "straight", "through", "u-turn", "uturn", "u turn",
    "approach", "direction", "inflow", "outflow", "in-", "out-",
)
LOCATION_HINTS = ("location", "intersection", "junction", "site", "chainage", "code")

TIME_RE = re.compile(r"^\s*\d{3,4}\s*-\s*\d{3,4}\s*$")          # 0800-0815 / 0800 - 0900
OD_RE = re.compile(r"^(.+?)\s+To\s+(.+?)$", re.IGNORECASE)      # "A To B"
# Decimal degrees: Jaipur is ~26.8N, 75.7E. Two such numbers together = a pin.
COORD_RE = re.compile(r"\b\d{1,3}\.\d{4,}\b")


def cell_repr(v):
    """Show the value as stored. No coercion, no formatting, None stays None."""
    if v is None:
        return ""
    return str(v)


def scan_strings(wb):
    """Every distinct string in the workbook, with where it was first seen."""
    seen = {}
    counts = Counter()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.strip():
                    s = v.strip()
                    counts[s] += 1
                    seen.setdefault(s, f"{ws.title}!{c.coordinate}")
    return counts, seen


def probe(path):
    print(BAR)
    print(f"DEEP PROBE : {path.name}")
    print(f"PATH       : {path}")
    print(BAR)

    # data_only=True gives the cached values Excel last wrote. The formulas
    # themselves are what the audit stage checks; here we want what the sheet shows.
    wb = load_workbook(path, data_only=True)

    # --- 1. sheet names -----------------------------------------------------
    print(f"\n[1] SHEET NAMES  ({len(wb.sheetnames)} sheets)")
    print(SUB)
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        print(f"  {i:>2}. {name:<22} dims={ws.dimensions:<12} "
              f"max_row={ws.max_row:<5} max_col={ws.max_column:<4} "
              f"merged={len(ws.merged_cells.ranges)}")

    # --- 2. first sheet, raw ------------------------------------------------
    ws = wb[wb.sheetnames[0]]
    print(f"\n[2] FIRST SHEET RAW  — {ws.title!r}")
    print(SUB)
    print(f"  dimensions : {ws.dimensions}")
    print(f"  rows x cols: {ws.max_row} x {ws.max_column}")
    merges = [str(r) for r in ws.merged_cells.ranges]
    print(f"  merged cell ranges ({len(merges)}): {', '.join(sorted(merges)) or 'none'}")
    print(f"\n  First {RAW_ROWS} rows verbatim. Blank = empty or merged-into "
          f"(only the top-left of a merge carries a value).\n")

    ncol = min(ws.max_column, 16)
    header = "  row | " + " | ".join(f"{get_column_letter(c):^18}" for c in range(1, ncol + 1))
    print(header)
    print("  " + "-" * (len(header) - 2))
    for r in range(1, min(RAW_ROWS, ws.max_row) + 1):
        cells = []
        for c in range(1, ncol + 1):
            txt = cell_repr(ws.cell(row=r, column=c).value)
            cells.append(f"{txt[:18]:^18}" if txt else " " * 18)
        line = f"  {r:>3} | " + " | ".join(cells)
        print(line.rstrip())

    # --- 3. location / coordinates -----------------------------------------
    print(f"\n[3] INTERSECTION NAME / LOCATION / COORDINATES")
    print(SUB)
    hits, coords = [], []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=12):
            for c in row:
                v = c.value
                if not isinstance(v, str):
                    continue
                low = v.lower()
                if any(h in low for h in LOCATION_HINTS):
                    # the label's value usually sits in the cell to its right
                    nxt = sheet.cell(row=c.row, column=c.column + 1).value
                    nxt2 = sheet.cell(row=c.row, column=c.column + 2).value
                    hits.append((sheet.title, c.coordinate, v.strip(),
                                 cell_repr(nxt) or cell_repr(nxt2)))
                if COORD_RE.search(v):
                    coords.append((sheet.title, c.coordinate, v.strip()))
    shown = set()
    for sh, coord, label, val in hits:
        key = (label, val)
        if key in shown:
            continue
        shown.add(key)
        print(f"  {sh + '!' + coord:<22} {label:<22} -> {val!r}")
    print(f"\n  Decimal-degree coordinates found: "
          f"{coords if coords else 'NONE — no lat/long anywhere in the workbook'}")
    if not coords:
        print("  >>> The workbooks carry no georeference. Junction pins must come")
        print("      from elsewhere (map pins, or the DWG once converted).")

    # --- 4/5/6. label discovery --------------------------------------------
    counts, seen = scan_strings(wb)
    veh, mov, times, ods = [], [], [], []
    for s in counts:
        low = s.lower()
        if TIME_RE.match(s):
            times.append(s)
        if any(h in low for h in VEHICLE_HINTS):
            veh.append(s)
        if any(h in low for h in MOVEMENT_HINTS):
            mov.append(s)
        if OD_RE.match(s) and " to " in low:
            ods.append(s)

    print(f"\n[4] VEHICLE CLASS LABELS  ({len(veh)} distinct)")
    print(SUB)
    for s in sorted(veh, key=lambda x: -counts[x]):
        print(f"  x{counts[s]:<5} {seen[s]:<18} {s!r}")

    print(f"\n[5] TIME / INTERVAL LABELS  ({len(times)} distinct)")
    print(SUB)
    q = sorted(t for t in times if "-" in t and " - " not in t)
    h = sorted(t for t in times if " - " in t)
    print(f"  15-min bins ({len(q)}): {q[0] if q else '-'} ... {q[-1] if q else '-'}")
    print(f"    full list: {', '.join(q)}" if q else "")
    print(f"\n  rolling hour windows ({len(h)}): {h[0] if h else '-'} ... {h[-1] if h else '-'}")
    print(f"    full list: {', '.join(h)}" if h else "")
    dates = set()
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=8):
            for c in row:
                if hasattr(c.value, "isoformat"):
                    dates.append if False else dates.add(str(c.value))
    print(f"\n  date/datetime cells in header rows: {sorted(dates) or 'none'}")

    print(f"\n[6] MOVEMENT LABELS AND ARM NAMES  ({len(mov)} distinct movement-ish)")
    print(SUB)
    for s in sorted(mov, key=lambda x: -counts[x]):
        print(f"  x{counts[s]:<5} {seen[s]:<18} {s!r}")
    print(f"\n  Origin->destination strings ({len(ods)}):")
    arms = set()
    for s in sorted(ods):
        m = OD_RE.match(s)
        arms.add(m.group(1).strip())
        arms.add(m.group(2).strip())
        print(f"    {s}")
    print(f"\n  Distinct arm names implied by those: {sorted(arms)}")
    print(f"  >>> {len(arms)} arms x {len(ods)} directed movements. "
          f"U-turn labels present: {any('u' == s.lower()[:1] and 'turn' in s.lower() for s in counts)}")


def sweep():
    """Sheet names and intersection identity for all 12 workbooks."""
    print(f"\n\n{BAR}")
    print("SWEEP — all workbooks")
    print(BAR)

    files = []
    for d in SURVEY_DIRS:
        files.extend(workbooks(SOURCE / d))
    print(f"{len(files)} workbooks found\n")

    signatures = defaultdict(list)
    for f in files:
        wb = load_workbook(f, data_only=True, read_only=True)
        names = wb.sheetnames
        ws = wb[names[0]]
        grid = {}
        for row in ws.iter_rows(min_row=1, max_row=10):
            for c in row:
                if c.value is not None:
                    grid[c.coordinate] = c.value
        wb.close()

        name = cell_repr(grid.get("B2"))
        code = cell_repr(grid.get("B3"))
        date = cell_repr(grid.get("I1"))
        day = cell_repr(grid.get("I2"))
        approaches = [cell_repr(grid.get(f"A{r}")) for r in (6, 7, 8, 9)]

        print(f"  {f.parent.name}/{f.name}")
        print(f"    location name : {name!r}   code: {code!r}")
        print(f"    date / day    : {date}  {day}")
        print(f"    approaches    : {approaches}")
        print(f"    sheets ({len(names)}): {', '.join(names)}")
        print()
        signatures[tuple(names)].append(f.name)

    print(SUB)
    print(f"Distinct sheet-name signatures across the {len(files)} files: {len(signatures)}")
    for sig, members in signatures.items():
        print(f"  {len(members)} files share a {len(sig)}-sheet layout: {sorted(set(members))}")


if __name__ == "__main__":
    if not PRIMARY.exists():
        raise SystemExit(f"Not found: {PRIMARY}")
    probe(PRIMARY)
    sweep()
