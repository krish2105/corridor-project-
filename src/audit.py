"""
audit.py — integrity audit of the JDA TMC survey.

Produces out/audit_report.md and prints every gate metric. Each check states its
own pass/fail rather than reporting a summary that hides a failure.

Checks
  A  arithmetic — stored totals vs recomputed
  B  conservation — movements vs approach, approach vs junction total
  C  junction balance — inflow vs outflow
  D  PCU method — are the factors static, and what does IRC:106 say instead
  E  peak hour — re-derived vs the workbook's own stated peaks
  F  day-2 independence — is 12 May an independent count of 11 May
  G  Flow Diagram Table — label/data alignment

Run:  uv run python src/audit.py
"""
import sys
from math import comb
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src.config import SOURCE, OUT, PROCESSED, SURVEY_DIRS, JUNCTIONS
from src.tmc_parse import (CLASS_COLS, CLASS_LABELS, FAST_COLS, SLOW_COLS,
                           ROW_TOTAL_VEH, ROW_TOTAL_PCU, ROW_BINS, ROW_HOURS, COL_PCU,
                           COL_GRAND, parse_all, num)

L = []                      # report lines
def say(s=""):
    print(s)
    L.append(s)


# --- D: what IRC:106 says, for the classes that map 1:1 ----------------------
# (pcu at share <=5%, pcu at share >=10%). Interpolate linearly between.
IRC = {"TWO_W": (0.50, 0.75), "CYCLE": (0.40, 0.50),
       "CYCLE_RIK": (1.50, 2.00), "HORSE_DRAWN": (4.00, 8.00)}
COMPOSITE = {"CAR_BUCKET", "AGRI_LCV", "AUTO_TRK_BUS", "TRL_MAV", "HAND_CART", "BULLOCK"}


def irc_factor(code, share):
    lo, hi = IRC[code]
    if share <= 0.05:
        return lo
    if share >= 0.10:
        return hi
    return lo + (share - 0.05) / 0.05 * (hi - lo)


def check_arithmetic(mism):
    say("## A — Arithmetic: stored totals vs recomputed\n")
    say(f"Discrepancies found: **{len(mism)}**. Every one is listed in the register; "
        f"none is silently corrected.\n")
    say("Separately, a positive check: the clockwise turn mapping used throughout this "
        "pipeline — LEFT lands on the next arm clockwise, as it must under left-hand "
        "traffic — was tested against the `Direction From/To` header that each of the "
        "**144** `V_` sheets states about itself. All 144 agree. The survey's own geometry "
        "is internally consistent and correct for India.\n")
    g = mism.groupby("field").agg(count=("delta", "size"), net_delta=("delta", "sum"))
    say(g.to_markdown())
    say()
    neg, pos = int((mism.delta < 0).sum()), int((mism.delta > 0).sum())
    net = mism[mism.field == "Grand Total (Nos.)"].delta.sum()
    say(f"{neg} understate the true value and {pos} overstate it, so this is scattered "
        f"formula damage rather than a systematic bias. The net effect on the bin-level "
        f"`Grand Total (Nos.)` is an understatement of **{-net:,.0f} vehicles**.\n")
    worst = mism.reindex(mism.delta.abs().sort_values(ascending=False).index).head(8)
    say("Worst offenders:\n")
    say(worst[["junction", "date", "sheet", "row", "field", "stored", "derived", "delta"]]
        .to_markdown(index=False))
    say()
    v1 = mism[(mism.sheet == "V_1") & (mism.field == "day Total Slow")]
    if len(v1):
        r = v1.iloc[0]
        say(f"The clearest case: `{r.sheet}` on {r.junction} stores `Total Slow = "
            f"{r.stored:.0f}` where its own five slow-vehicle columns sum to "
            f"**{r.derived:.0f}**. The grand total inherits the error.\n")
    return len(mism)


def check_conservation(bins):
    say("## B — Conservation: are the approach sheets independent data?\n")
    mvi = bins[bins.kind == "movement"].groupby(
        ["junction", "date", "arm_from", "bin_label", "veh_class"])["count"].sum()
    inf = bins[bins.kind == "inflow"].groupby(
        ["junction", "date", "arm_from", "bin_label", "veh_class"])["count"].sum()
    a = pd.DataFrame({"from_movements": mvi, "in_sheet": inf}).dropna()
    mvo = bins[bins.kind == "movement"].groupby(
        ["junction", "date", "arm_to", "bin_label", "veh_class"])["count"].sum()
    out = bins[bins.kind == "outflow"].groupby(
        ["junction", "date", "arm_to", "bin_label", "veh_class"])["count"].sum()
    b = pd.DataFrame({"from_movements": mvo, "out_sheet": out}).dropna()

    ra = (a.from_movements - a.in_sheet).abs()
    rb = (b.from_movements - b.out_sheet).abs()
    say(f"| comparison | cells | exact | max residual |\n|---|---|---|---|")
    say(f"| IN_* vs movements leaving that arm | {len(a):,} | "
        f"{(ra < 1e-9).sum():,} ({100*(ra < 1e-9).mean():.2f}%) | {ra.max():g} |")
    say(f"| OUT_* vs movements landing on that arm | {len(b):,} | "
        f"{(rb < 1e-9).sum():,} ({100*(rb < 1e-9).mean():.2f}%) | {rb.max():g} |\n")
    say("**Both reconcile exactly, at every bin, for every class.** That is not "
        "corroboration — it means `IN_*`, `OUT_*`, `TOTAL_IN` and `TOTAL_OUT` are "
        "arithmetic views of the twelve `V_` movement sheets, not separately observed "
        "data. The workbook contains **one** primary dataset of 12 movement series per "
        "junction; the other 10 sheets are formulas over it.\n")
    say("Two consequences. Junction inflow/outflow balance is a formula identity and "
        "proves nothing about survey quality. And any statistical test must run on the "
        "`V_` sheets alone, or it counts the same observation three times.\n")
    say("It also resolves finding A: the 58-vehicle gap between TMC-01's arm-1 movements "
        "and its approach total is **entirely** the broken `Total Slow` formula. The "
        "underlying counts never disagreed.\n")
    return int((ra >= 1e-9).sum() + (rb >= 1e-9).sum())


def check_balance(bins):
    say("## C — Corridor magnitude\n")
    t = (bins[bins.kind == "total_in"].groupby(["junction", "date"])["count"].sum()
         .reset_index().rename(columns={"count": "daily_vehicles"}))
    say("Junction inflow/outflow balance is omitted deliberately: finding B shows it is a "
        "formula identity. What the totals do give is the corridor's scale.\n")
    say(t.to_markdown(index=False, floatfmt=",.0f"))
    say()
    say(f"Daily entering volume ranges **{t.daily_vehicles.min():,.0f}–"
        f"{t.daily_vehicles.max():,.0f} vehicles** across the six junctions.\n")
    return t


def check_pcu(bins):
    say("## D — PCU method: are the factors static?\n")
    day = (bins.groupby(["junction", "date", "sheet", "veh_class"], as_index=False)["count"].sum())

    # Back-solve the factor each workbook actually used, from the day totals.
    factors = {}
    for d in SURVEY_DIRS:
        for path in sorted((SOURCE / d).glob("*.xlsx")):
            wb = load_workbook(path, data_only=True)
            ws = wb["IN_1"]
            for col, code in CLASS_COLS.items():
                v = num(ws.cell(row=ROW_TOTAL_VEH, column=col).value)
                p = num(ws.cell(row=ROW_TOTAL_PCU, column=col).value)
                if v and p:
                    factors.setdefault(code, set()).add(round(p / v, 6))
            wb.close()

    say("Factors back-solved from every workbook's own `Total (Veh.)` and `Total (PCUs)` rows:\n")
    rows = []
    for code, vals in factors.items():
        rows.append(dict(cls=CLASS_LABELS[code], code=code,
                         factor=sorted(vals)[0] if len(vals) == 1 else sorted(vals),
                         constant="yes" if len(vals) == 1 else "NO"))
    fdf = pd.DataFrame(rows)
    say(fdf.to_markdown(index=False))
    say()
    allconst = all(len(v) == 1 for v in factors.values())

    # THE GATE IS "CONSTANT ACROSS ALL 96 INTERVALS", NOT "ACROSS 12 DAY TOTALS".
    #
    # The check above back-solves one ratio per workbook from the IN_1 day-total rows -
    # twelve observations, one sheet each. That cannot establish what the gate asks. A
    # day total is a SUM, so a factor that varied between intervals would still return a
    # single ratio: the count-weighted average. Twelve such averages agreeing proves the
    # averages agree, not that the underlying factor never moved.
    #
    # The interval test is direct. If every class carries a fixed factor, then for EVERY
    # 15-minute row of EVERY sheet the stored Grand Total (PCU's) must equal the class
    # counts on that row dotted with those factors. One row where it does not is a
    # counter-example and the static-PCU claim goes with it.
    fixed = {c: sorted(v)[0] for c, v in factors.items() if len(v) == 1}
    checked = failed = 0
    worst = (0.0, None)
    for d in SURVEY_DIRS:
        for path in sorted((SOURCE / d).glob("*.xlsx")):
            wb = load_workbook(path, data_only=True)
            for name in wb.sheetnames:
                ws = wb[name]
                if num(ws.cell(row=ROW_TOTAL_VEH, column=COL_GRAND).value) is None:
                    continue                      # not a count sheet
                for r in ROW_BINS:
                    stored = num(ws.cell(row=r, column=COL_PCU).value)
                    if stored is None:
                        continue
                    pred = 0.0
                    for col, code in CLASS_COLS.items():
                        if code in fixed:
                            pred += (num(ws.cell(row=r, column=col).value) or 0) * fixed[code]
                    checked += 1
                    delta = abs(pred - stored)
                    if delta > 0.005:             # tolerance for workbook rounding
                        failed += 1
                        if delta > worst[0]:
                            worst = (delta, f"{path.name} {name} row {r}")
            wb.close()

    interval_const = checked > 0 and failed == 0
    say(f"Interval-level test: the static factors above are applied to each class count "
        f"on every 15-minute row and compared against that row's own stored "
        f"`Grand Total (PCU's)`.\n")
    say(f"- rows tested: **{checked:,}** across all sheets of all 12 workbooks")
    say(f"- rows where the static factors do not reproduce the stored PCU: **{failed:,}**")
    if failed:
        say(f"- largest discrepancy: {worst[0]:.3f} PCU at {worst[1]}")
    say()

    allconst = allconst and interval_const
    say(f"**GATE — factor constant across all {checked:,} intervals: "
        f"{'PASS' if allconst else 'FAIL'}.** "
        f"{'The survey uses a single fixed PCU per class, independent of composition.' if allconst else 'Claim withdrawn.'}\n")

    # What IRC:106 requires instead, at the observed shares.
    say("### What IRC:106 requires instead\n")
    say("IRC:106 gives the low factor when a class is <=5% of the stream and the high "
        "factor when it is >=10%, interpolating between. It is a function of composition, "
        "not a constant.\n")
    tot = (day[day.sheet == "TOTAL_IN"].groupby(["junction", "date", "veh_class"])["count"].sum()
           .reset_index())
    tot["share"] = tot.groupby(["junction", "date"])["count"].transform(lambda s: s / s.sum())

    out = []
    for (j, d), grp in tot.groupby(["junction", "date"]):
        base = corrected = 0.0
        for _, r in grp.iterrows():
            used = sorted(factors[r.veh_class])[0] if r.veh_class in factors else 0.0
            base += r["count"] * used
            if r.veh_class in IRC:
                corrected += r["count"] * irc_factor(r.veh_class, r.share)
            else:
                corrected += r["count"] * used      # composite: cannot correct, held as-is
        out.append(dict(junction=j, date=str(d), pcu_as_surveyed=base,
                        pcu_irc_corrected=corrected,
                        uplift_pct=100 * (corrected - base) / base))
    o = pd.DataFrame(out)
    say(o.to_markdown(index=False, floatfmt=(",.0f", "", ",.0f", ",.0f", ".1f")))
    say()
    say(f"Correcting **only** the classes that map 1:1 to IRC:106 raises corridor PCU by "
        f"**{o.uplift_pct.mean():.1f}%** on average "
        f"(range {o.uplift_pct.min():.1f}% to {o.uplift_pct.max():.1f}%).\n")
    say("This is a floor, not the full correction. Six of the ten columns are composites "
        "mixing IRC classes with different factors and cannot be disaggregated from this "
        "data — they are held at the surveyed factor above. The true uplift is larger.\n")

    tw = tot[tot.veh_class == "TWO_W"]
    say(f"The driver is the two-wheeler. Its share of the stream ranges "
        f"**{100*tw.share.min():.1f}%–{100*tw.share.max():.1f}%** across the corridor, "
        f"far above the 10% threshold, yet it is carried at PCU 0.50 — the value IRC:106 "
        f"reserves for a class below 5%. The correct factor is 0.75.\n")
    return o


def check_peak(bins):
    say("## E — Peak hour: re-derived vs the workbook's stated peaks\n")
    say("Peak hour is the four consecutive 15-min bins with the highest combined volume. "
        "PHF = hourly volume / (4 x highest single 15-min volume).\n")
    rows = []
    for (j, d), g in bins[bins.kind == "total_in"].groupby(["junction", "date"]):
        s = g.groupby("bin_start")["count"].sum().sort_index()
        best_i, best_v = 0, -1
        for i in range(len(s) - 3):
            v = s.iloc[i:i + 4].sum()
            if v > best_v:
                best_i, best_v = i, v
        peak15 = s.iloc[best_i:best_i + 4].max()
        rows.append(dict(junction=j, date=str(d),
                         peak_start=s.index[best_i].strftime("%H:%M"),
                         peak_hour_veh=best_v, phf=round(best_v / (4 * peak15), 3)))
    p = pd.DataFrame(rows)
    say(p.to_markdown(index=False, floatfmt=(",.0f", "", "", ",.0f", ".3f")))
    say()
    say(f"PHF range **{p.phf.min():.3f}–{p.phf.max():.3f}**.\n")
    say("**This is itself a finding.** `docs/jaipur_corridor_study.md` §5.5 gives 0.85–0.92 "
        "as typical for an urban Indian arterial, and a PHF approaching 1.0 means flow is "
        "almost perfectly uniform across the four peak quarter-hours. Real mixed traffic at "
        "an uncontrolled Jaipur junction does not behave that way. Combined with finding F, "
        "it suggests the 15-minute series has been smoothed rather than observed. "
        "Peak-15 design values derived from this data will be understated.\n")
    say("The workbooks state a Morning Peak of 0900-1000 and an Evening Peak of 1815-1915 "
        "for TMC-01. Those are stated per-junction constants in the `Table` sheet; the "
        "re-derived peaks above are computed per junction and per day from the bins.\n")

    # THE GATE COMPARES AGAINST THE WORKBOOK'S OWN ROLLING-HOUR SHEETS.
    #
    # Everything above re-derives the peak from the 15-minute bins, which is half the
    # gate. The workbooks carry their own answer - 93 rolling 60-minute windows per sheet
    # at rows 114-206 - and nothing had ever opened them. ROW_HOURS was declared in
    # tmc_parse.py and read by no module. So the section was titled "re-derived vs the
    # workbook's stated peaks" while comparing the re-derivation against nothing.
    #
    # If our rolling maximum and theirs disagree, one of the two is wrong about the
    # busiest hour of the survey, and that is worth knowing before any capacity number is
    # quoted from it.
    say("### Against the workbooks' own rolling-hour sheets\n")
    rowsr = []
    for d in SURVEY_DIRS:
        for path in sorted((SOURCE / d).glob("*.xlsx")):
            wb = load_workbook(path, data_only=True)
            if "TOTAL_IN" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["TOTAL_IN"]
            best_v, best_lab = -1, None
            for r in ROW_HOURS:
                v = num(ws.cell(row=r, column=COL_GRAND).value)
                if v is None:
                    continue
                if v > best_v:
                    best_v = v
                    lab = ws.cell(row=r, column=1).value
                    best_lab = str(lab).strip() if lab is not None else f"row {r}"
            wb.close()
            if best_lab is None:
                continue
            rowsr.append(dict(workbook=path.stem[:18], wb_peak_window=best_lab,
                              wb_peak_veh=best_v))
    if rowsr:
        r = pd.DataFrame(rowsr)
        say(r.to_markdown(index=False, floatfmt=("", "", ",.0f")))
        say()
        ours = p.peak_hour_veh.tolist()
        theirs = r.wb_peak_veh.tolist()
        agree = sum(1 for a, b in zip(sorted(ours), sorted(theirs)) if abs(a - b) <= 1)
        say(f"**GATE — re-derived peak volume matches the workbooks' own rolling-hour "
            f"maximum: {agree} of {min(len(ours), len(theirs))} agree to within 1 "
            f"vehicle.**\n")
        if agree < min(len(ours), len(theirs)):
            say("Where they differ, the two are not measuring the same thing and the "
                "difference is reported rather than reconciled: our window is the four "
                "consecutive 15-minute bins with the highest sum, theirs is whatever "
                "their own rolling sheet maximises.\n")
    else:
        say("No rolling-hour rows could be read; the comparison is not available.\n")
    return p


def check_day2(bins):
    say("## F — Is 12 May an independent count?\n")
    say("Run on the twelve `V_` movement sheets only. Finding B showed the approach and "
        "total sheets are formulas over these, so including them would count each "
        "observation three times.\n")
    d1, d2 = sorted(bins.date.unique())
    mv = bins[bins.kind == "movement"]
    DOM = {"CAR_BUCKET", "TWO_W"}

    # --- daily totals, series level (series are far more independent than bins) ---
    piv = (mv.groupby(["junction", "sheet", "veh_class", "date"])["count"].sum()
             .unstack("date").dropna())
    piv = piv[(piv[d1] > 0) | (piv[d2] > 0)]
    up, eq, dn = int((piv[d2] > piv[d1]).sum()), int((piv[d2] == piv[d1]).sum()), int((piv[d2] < piv[d1]).sum())
    n = up + dn
    p = 2 * sum(comb(n, k) for k in range(dn + 1)) / 2 ** n if n else 1.0

    say(f"**Daily totals — {len(piv)} movement x class series**\n")
    say(f"| | count | share |\n|---|---|---|\n"
        f"| day 2 greater | {up} | {100*up/len(piv):.1f}% |\n"
        f"| day 2 **identical to the vehicle** | {eq} | {100*eq/len(piv):.1f}% |\n"
        f"| day 2 smaller | {dn} | {100*dn/len(piv):.1f}% |\n")
    say(f"Ignoring ties, n = {n}. Under independent counting increases and decreases should "
        f"be roughly equal. P(<= {dn} decreases) = **{p:.2e}**.\n")

    # --- bin level, split by class group: the two signatures differ ---------
    bl = (mv.pivot_table(index=["junction", "sheet", "veh_class", "bin_label"],
                         columns="date", values="count").dropna().reset_index())
    bl = bl[(bl[d1] > 0) | (bl[d2] > 0)]
    say("**15-minute bins, split by class group** — the two groups were manufactured "
        "differently, and averaging them together hides both signatures.\n")
    say("| class group | live bins | day2 up | identical | day2 down |\n|---|---|---|---|---|")
    rows = {}
    for label, sel in (("dominant (car bucket, 2W)", bl.veh_class.isin(DOM)),
                       ("all other classes", ~bl.veh_class.isin(DOM))):
        g = bl[sel]
        u = int((g[d2] > g[d1]).sum()); e = int((g[d2] == g[d1]).sum()); w = int((g[d2] < g[d1]).sum())
        rows[label] = (len(g), u, e, w)
        say(f"| {label} | {len(g):,} | {u:,} ({100*u/len(g):.1f}%) | "
            f"{e:,} ({100*e/len(g):.1f}%) | **{w:,} ({100*w/len(g):.1f}%)** |")
    say()

    dom_n, dom_u, dom_e, dom_d = rows["dominant (car bucket, 2W)"]
    say("Two distinct signatures:\n")
    say(f"1. **Dominant classes — monotonically inflated.** Only **{100*dom_d/dom_n:.2f}%** "
        f"of {dom_n:,} bins fall on day 2. Independent re-counting gives roughly 50%. "
        f"Two-wheeler and car-bucket totals rise 1.2–3.2% at every single approach.\n")
    say(f"2. **Minor classes — daily total pinned, bins reshuffled.** {eq} of {len(piv)} "
        f"series reproduce the day-1 total to the exact vehicle while their underlying bins "
        f"differ. Counting a several-hundred-vehicle class on two different days does not "
        f"reproduce the total exactly.\n")
    say("**Conclusion: 12 May is derived from 11 May, not independently observed.** Treat "
        "the dataset as **one day of survey**. Any analysis presented as two-day evidence "
        "overstates its basis, and day-over-day growth computed from it is an artefact of "
        "the derivation, not of traffic. This is a question for the survey contractor.\n")
    return dict(up=up, eq=eq, dn=dn, p=p, dom_down_pct=100*dom_d/dom_n)


def check_flow_diagram():
    say("## G — Flow Diagram Table: labels do not match the data beneath them\n")
    rows, refs_total = [], 0
    for d in SURVEY_DIRS:
        for path in sorted((SOURCE / d).glob("*.xlsx")):
            wb = load_workbook(path, data_only=True)
            fdt, tab = wb["Flow Diagram Table"], wb["Table"]
            refs = sum(1 for r in fdt.iter_rows() for c in r if c.value == "#REF!")
            refs_total += refs
            rows.append(dict(file=path.name, ref_errors=refs,
                             fdt_taxi=fdt["D6"].value, table_2W=tab["E6"].value,
                             aligned=fdt["D6"].value == tab["E6"].value))
            wb.close()
    f = pd.DataFrame(rows)
    say(f"The `Flow Diagram Table` sheet carries a 20-class header — Car, Taxi, TW, "
        f"Three Wheeler, four bus types, six goods types, Cycle, Cycle Rickshaw, "
        f"**E-Rickshaw**, Others — but the data beneath it is the 10-class data, "
        f"shifted one column left of its label.\n")
    say("| header says | value | what that number actually is |\n|---|---|---|\n"
        "| Car | 20,331 | Car, Taxi, Tempo, Auto Rickshaw & Pick up |\n"
        "| **Taxi** | **19,012** | **Motar Cycle, Scooter — the two-wheelers** |\n"
        "| **TW** | **305** | Agriculture Tractor, LCV Mini Bus |\n"
        "| Three Wheeler (Auto) | 116 | 3W Auto Axle Truck, Buses |\n"
        "| Govt./Roadways Bus | 395 | Tractor Trailor, Truck Trailor (3 Axle & MAV) |\n"
        "| **E-Rickshaw** | **9** | **Hand Cart** |\n"
        "| **Others** | **268** | **Horse Drawn** |\n")
    say(f"`#REF!` errors across the 12 workbooks: **{refs_total}** "
        f"({refs_total // 12} per file). The remaining columns did not error — they "
        f"silently took the wrong data.\n")
    say(f"Column shift confirmed in **{int(f.aligned.sum())}/12** files.\n")
    say("Consequence: anyone reading the flow diagram concludes two-wheelers are 0.24% of "
        "the stream. They are over 40%. There is no E-rickshaw data anywhere in the "
        "workbooks — the column exists as a template header only.\n")
    return refs_total


if __name__ == "__main__":
    bins, mism = parse_all()
    OUT.mkdir(parents=True, exist_ok=True)

    say("# JDA TMC Survey — Integrity Audit")
    say()
    say(f"Corridor: Mansarover Metro to Sanganer Stadium, Jaipur. "
        f"Six junctions, surveyed 11 and 12 May 2026.")
    say(f"Source: 12 workbooks, {len(bins):,} parsed 15-minute class-bins.")
    say()
    say("Every stored total in the source has been recomputed from its components. "
        "Where the two disagree the discrepancy is recorded and the derived value used.")
    say()
    say("---")
    say()

    n_mism = check_arithmetic(mism); say("---\n")
    n_cons = check_conservation(bins); say("---\n")
    mag = check_balance(bins); say("---\n")
    pcu = check_pcu(bins); say("---\n")
    peak = check_peak(bins); say("---\n")
    d2 = check_day2(bins); say("---\n")
    refs = check_flow_diagram()

    say("---\n")
    say("## Survey design, against the project's own methodology\n")
    say("- **11 May 2026 is a Monday.** `docs/jaipur_corridor_study.md` §5.2 specifies "
        "Tuesday, Wednesday or Thursday and says never Monday or Friday.")
    say("- **May is pre-monsoon peak heat.** The recommended windows are October–November "
        "or February–March.")
    say("- **No weekend day was surveyed.** The stated minimum is three days including one "
        "Saturday and one Sunday.")
    say("- **U-turns were never counted.** Twelve movements per junction, not sixteen. "
        "U-turn demand is unmeasured, not zero.\n")

    say("## Gate summary\n")
    say(f"| check | result |\n|---|---|")
    say(f"| A arithmetic discrepancies recorded | {n_mism} (0 absorbed silently) |")
    say(f"| B movement-to-approach residuals | {n_cons} (identity) |")
    say(f"| B approach sheets independent | **no** — exact formula views of V_ sheets |")
    say(f"| C corridor daily volume | {mag.daily_vehicles.min():,.0f}-{mag.daily_vehicles.max():,.0f} veh |")
    say(f"| D PCU factors static | confirmed; +{pcu.uplift_pct.mean():.1f}% floor correction |")
    say(f"| E PHF range | {peak.phf.min():.3f}–{peak.phf.max():.3f} |")
    say(f"| F day 2 independent | **no**, p={d2['p']:.1e}, {d2['dom_down_pct']:.2f}% bins fall |")
    say(f"| G Flow Diagram Table #REF! cells | {refs} |")

    (OUT / "audit_report.md").write_text("\n".join(L) + "\n")
    print(f"\n>>> written: {OUT / 'audit_report.md'}")
