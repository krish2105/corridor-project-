"""
junction_books.py — twelve editable Excels: 6 junctions x 2 days, 12 movement sheets each.

The issued survey is twelve workbooks - six junctions counted on two days - and the
deliverable mirrors that one for one, so a reviewer can lay our J3 Day-1 book beside the
issued 04_TMC (11-05-2026) file and reconcile cell against cell. Day-2 books carry the
audit's warning on every sheet: 396 of 555 series reproduce day 1 exactly, so that day is
derived, not independently observed.

Each book also carries a U-TURN BAYS sheet - the updated movement set under the signal-free
scheme: which three movements feed each bay, the demand, the conflicting flow, and the
gap-acceptance verdict, straight from scheme_test.

WHAT THE REVIEWER ASKED FOR
Per junction, every one of the twelve surveyed movements on its own sheet, grouped the
way the movements group: four LEFT sheets, four STRAIGHT sheets, four RIGHT sheets.
Each sheet carries a turning-movement diagram, the movement's full vehicle-wise
distribution (96 fifteen-minute bins by ten classes, exactly as parsed from the issued
workbook), the class totals with shares, and an editable chart. Nothing synthetic: every
number is re-derived from the source cells, and the gate below reconciles each sheet's
total against the parse.

Run:  uv run python src/junction_books.py
"""
import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.path import Path as MplPath
import matplotlib.patches as mpatches

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src.config import JUNCTIONS, JUNCTION_COORDS, OUT, SCHEME_LABEL
from src.routes import route as scheme_route
from src.spelling import fix as spell
from src.tmc_parse import CLASS_LABELS, parse_all

BOOKS = OUT / "junction_books"
IMGS = BOOKS / "_diagrams"

INK, MUTED, ACCENT = "14181A", "5C6663", "1B3A6B"
OKC, DEFECT, SUNK = "2C6249", "9E2B25", "F1F2ED"
TURN_HEX = {"Left": "#2C6249", "Straight": "#1B3A6B", "Right": "#9E2B25"}

H = Font(bold=True, size=14, color=INK)
SUBF = Font(size=9, color=MUTED)
TH = Font(bold=True, size=9, color=MUTED)
BOLD = Font(bold=True, size=10)
FILL = PatternFill("solid", start_color=SUNK)
THIN = Border(bottom=Side(style="thin", color="D5D9D4"))

# The sheet order the reviewer asked for: 4-4-4, movements grouped by turn.
GROUPS = [("Left", "L"), ("Straight", "S"), ("Right", "R")]


def diagram(arms, fi, ti, turn, path, r, bays, width_m):
    """
    The movement in its full context, dimensioned — not a lone arrow.

    Every diagram now carries: the measured carriageway width on the corridor; BOTH
    U-turn bays with their measured one-way distance from the junction (or "beyond
    drawing" where the CAD ends); and, for a movement the scheme bans, the complete
    re-route drawn leg by leg in amber beside the direct movement it replaces. A single
    arrow answered "which movement is this?"; a reviewer's question is "what does a
    driver DO?", and that needs the whole picture with numbers on it.

    Distances are the published ones from scheme_test - real chainage differences, never
    invented - and the schematic is not to scale, so every distance is written on it.
    """
    import math
    fig, ax = plt.subplots(figsize=(4.6, 4.6), dpi=120)
    ax.set_xlim(-1.55, 1.55); ax.set_ylim(-1.62, 1.5)
    ax.set_aspect("equal"); ax.axis("off")
    LANE = 0.14

    for i in range(4):
        th = math.pi / 2 * i
        ox, oy = math.sin(th), math.cos(th)
        wide = 30 if i in (0, 2) else 22          # corridor drawn wider than cross
        ax.plot([ox * .12, ox * 1.28], [oy * .12, oy * 1.28],
                color="#E2E4DF", lw=wide, solid_capstyle="butt", zorder=1)
        ax.text(ox * 1.42, oy * 1.4 - (0.03 if i in (1, 3) else 0),
                spell(arms[i]), ha="center", va="center", fontsize=7, color="#5C6663")
    # median line on the corridor
    ax.plot([0, 0], [-1.28, 1.28], color="#C9CDC6", lw=1.1, ls=(0, (4, 3)), zorder=2)

    def node(i, entering):
        th = math.pi / 2 * i
        ox, oy = math.sin(th), math.cos(th)
        px, py = math.cos(th), -math.sin(th)
        s = 1 if entering else -1
        return (ox * .5 + px * s * LANE, oy * .5 + py * s * LANE)

    def bez(fi_, ti_, colour, lw, dashed=False):
        x1, y1 = node(fi_, True); x2, y2 = node(ti_, False)
        off = (ti_ - fi_) % 4
        if off == 2:
            pth = MplPath([(x1, y1), (x2, y2)], [MplPath.MOVETO, MplPath.LINETO])
        else:
            ta = math.pi / 2 * fi_; tb = math.pi / 2 * ti_
            ax_, ay_ = -math.sin(ta), -math.cos(ta)
            bx_, by_ = math.sin(tb), math.cos(tb)
            det = ax_ * -by_ - ay_ * -bx_
            tt = ((x2 - x1) * -by_ - (y2 - y1) * -bx_) / det
            pth = MplPath([(x1, y1), (x1 + tt * ax_, y1 + tt * ay_), (x2, y2)],
                          [MplPath.MOVETO, MplPath.CURVE3, MplPath.CURVE3])
        ax.add_patch(mpatches.FancyArrowPatch(
            path=pth, arrowstyle="-|>", mutation_scale=13, lw=lw, color=colour,
            linestyle=(0, (5, 3)) if dashed else "solid", zorder=4))

    # BOTH bays, always: the back-to-back context every sheet was missing
    for side, ybay in (("north", 1.06), ("south", -1.06)):
        b = bays.get(side)
        ax.add_patch(mpatches.FancyArrowPatch(
            (LANE, ybay), (-LANE, ybay),
            connectionstyle=f"arc3,rad={-0.9 if side == 'north' else 0.9}",
            arrowstyle="-|>", mutation_scale=9, lw=1.6, color="#82600F", zorder=3))
        lab = ("beyond drawing" if not b or b.get("one_way_m") is None
               else f"{b['one_way_m']:,} m out")
        ax.text(0.30, ybay, f"U-bay {side}\n{lab}", fontsize=6.2,
                color="#82600F", va="center")

    amber = "#B8860B"
    if r["permitted"] == "re-routed":
        # ghost of the banned direct movement, then the real route in amber
        bez(fi, ti, TURN_HEX[turn], 1.4, dashed=True)
        corridor = fi in (0, 2)
        headI = (fi + 2) % 4 if corridor else (fi + 1) % 4
        # after the U-turn the driver re-enters the junction FROM the arm they ran out
        # toward - they went out toward headI, turned, and come back in on that same arm
        ybay = 1.06 if headI == 0 else -1.06
        sgn = 1 if headI == 0 else -1
        if not corridor:
            bez(fi, headI, amber, 2.6)                 # the forced left
        hx, hy = node(headI, False)
        if corridor:
            x1, y1 = node(fi, True)
            ax.plot([x1, hx], [y1, hy], color=amber, lw=2.6, zorder=4)
        ax.plot([hx, hx], [hy, ybay * .97], color=amber, lw=2.6, zorder=4)
        ax.add_patch(mpatches.FancyArrowPatch(
            (hx, ybay * .97), (-hx, ybay * .97),
            connectionstyle=f"arc3,rad={sgn * 0.9}",
            arrowstyle="-", lw=2.6, color=amber, zorder=4))
        bx2, by2 = node(headI, True)
        ax.plot([-hx, bx2], [ybay * .97, by2], color=amber, lw=2.6, zorder=4)
        bez(headI, ti, amber, 2.6)
        b = bays.get("north" if headI == 0 else "south") or {}
        note = ("route: no opening within the drawing on this side"
                if b.get("one_way_m") is None else
                f"route: {b['one_way_m']:,} m to the bay, "
                f"{b['detour_m']:,} m round trip")
        ax.text(0, -1.55, note, ha="center", fontsize=7, color=amber)
    else:
        bez(fi, ti, TURN_HEX[turn], 3.0)
        ax.text(0, -1.55, "permitted at the junction — unaffected by the scheme",
                ha="center", fontsize=7, color="#2C6249")

    # measured carriageway dimension, provisional
    if width_m:
        ax.text(-1.5, 1.42, f"carriageway {2 * width_m:.1f} m both directions\n"
                            f"({width_m:.1f} m/dir · CAD-derived, provisional)",
                ha="left", va="top", fontsize=6.4, color="#5C6663")

    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def sheet_for(wb, code, arms, fi, ti, turn, day_mv, day_label, derived, img_path, bays):
    frm, to = arms[fi], arms[ti]
    title = f"{turn[0]}{fi + 1} {spell(frm)[:14]}"
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False

    r = scheme_route(fi, ti)
    ws["A1"] = (f"{SCHEME_LABEL[code]} — {spell(frm)} → {spell(to)}  ({turn})")
    ws["A1"].font = H
    ws["A2"] = (f"Survey sheet {code} · movement V_{fi * 3 + [g for g, _ in GROUPS].index(turn) + 1} "
                f"· India drives on the left; the left turn is the next arm clockwise")
    ws["A2"].font = SUBF
    if r["permitted"] == "re-routed":
        b = bays.get(r["bay"]) or {}
        dist = ("no opening within the drawing on this side"
                if b.get("one_way_m") is None else
                f"{b['one_way_m']:,} m to the bay, {b['detour_m']:,} m round trip"
                + (" — and that opening is a junction mouth"
                   if b.get("bay_is_junction_mouth") else ""))
        ws["A3"] = ("BANNED at the junction under the signal-free scheme. Route: "
                    + " → ".join(r["legs"]) + f".  Measured: {dist}.")
    else:
        ws["A3"] = ("Under the signal-free scheme: unaffected — " + r["legs"][0])
    ws["A3"].font = Font(size=9, italic=True,
                         color=DEFECT if r["permitted"] == "re-routed" else OKC)

    img = XLImage(str(img_path)); img.anchor = "L1"
    ws.add_image(img)

    if derived:
        ws["A4"] = ("CAUTION: the audit found this day is derived from day 1 on most "
                    "series, not independently observed. Treat as a copy check, not data.")
        ws["A4"].font = Font(size=9, bold=True, color=DEFECT)

    # data block: 96 bins x classes for THIS day
    d1 = day_mv[(day_mv.arm_from == frm) & (day_mv.arm_to == to)]
    piv = d1.pivot_table(index="bin_label", columns="veh_class",
                         values="count", aggfunc="sum").fillna(0)
    classes = [c for c in CLASS_LABELS if c in piv.columns]
    piv = piv[classes]

    hdr = 5
    ws.cell(hdr, 1, "15-min bin").font = TH
    for j, c in enumerate(classes, start=2):
        cell = ws.cell(hdr, j, spell(CLASS_LABELS[c]))
        cell.font = TH; cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = FILL
    ws.cell(hdr, len(classes) + 2, "Total").font = TH
    ws.cell(hdr, 1).fill = FILL; ws.cell(hdr, len(classes) + 2).fill = FILL

    for i, (lab, row) in enumerate(piv.iterrows(), start=hdr + 1):
        ws.cell(i, 1, lab).font = Font(size=9)
        for j, c in enumerate(classes, start=2):
            ws.cell(i, j, int(row[c])).font = Font(size=9)
        last = get_column_letter(len(classes) + 1)
        ws.cell(i, len(classes) + 2, f"=SUM(B{i}:{last}{i})").font = Font(size=9)

    tot_r = hdr + len(piv) + 1
    ws.cell(tot_r, 1, f"Total ({day_label})").font = BOLD
    for j in range(2, len(classes) + 3):
        col = get_column_letter(j)
        c = ws.cell(tot_r, j, f"=SUM({col}{hdr + 1}:{col}{tot_r - 1})")
        c.font = BOLD; c.border = THIN

    day1 = d1.groupby("veh_class")["count"].sum()
    grand = day1.sum()
    ws.cell(tot_r + 1, 1, "Share of movement %").font = TH
    for j, c in enumerate(classes, start=2):
        ws.cell(tot_r + 1, j,
                round(100 * day1.get(c, 0) / grand, 2) if grand else 0).font = Font(size=9)

    ch = BarChart(); ch.type = "col"; ch.title = f"Vehicle-wise distribution, {day_label}"
    ch.height, ch.width = 7, 16; ch.legend = None
    ch.add_data(Reference(ws, min_col=2, max_col=len(classes) + 1,
                          min_row=hdr, max_row=tot_r), from_rows=False, titles_from_data=True)
    # one series: the totals row
    ch.add_data(Reference(ws, min_col=2, max_col=len(classes) + 1,
                          min_row=tot_r, max_row=tot_r), from_rows=True)
    ch.series = ch.series[-1:]
    ch.set_categories(Reference(ws, min_col=2, max_col=len(classes) + 1,
                                min_row=hdr, max_row=hdr))
    ws.add_chart(ch, f"A{tot_r + 5}")

    ws.column_dimensions["A"].width = 15
    for j in range(2, len(classes) + 3):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.freeze_panes = f"A{hdr + 1}"
    return int(grand)


def uturn_sheet(wb, code):
    """
    The updated movement set under the signal-free scheme, for this junction.

    Read from scheme_test's published output rather than recomputed, so the workbook and
    the dashboard cannot disagree about a verdict.
    """
    import json
    from src.routes import bay_movements, conflicting_direction
    p = OUT / "data" / "scheme_test.json"
    ws = wb.create_sheet("U-Turn Bays")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{SCHEME_LABEL[code]} — U-turn bays under the signal-free scheme"
    ws["A1"].font = H
    ws["A2"] = ("Back-to-back arrangement: one bay each side of the junction, as run on "
                "Noida's Dadri (DSC) Road. Each bay is fed by THREE movements, not one.")
    ws["A2"].font = SUBF
    rows = []
    if p.exists():
        rows = [u for u in json.loads(p.read_text())["uturns"]
                if u["junction"] == code]
    hdr = 4
    heads = ["Bay side", "Rejoins", "Movements it serves", "Bay distance m (one-way)",
             "Detour m (round trip)", "Demand veh/h",
             "Conflicting flow veh/h", "Critical gap s", "Capacity veh/h", "Verdict"]
    for j, htxt in enumerate(heads, start=1):
        c = ws.cell(hdr, j, htxt); c.font = TH; c.fill = FILL
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for i, u in enumerate(sorted(rows, key=lambda x: x["bay"]), start=hdr + 1):
        side = u["bay"]
        serves = ", ".join(f"{m['from_arm']}→{m['to_arm']}" for m in bay_movements(side))
        vc = u["vc_optimistic"]
        verdict = ("no viable gaps" if vc >= 3.0 else
                   f"fails, v/c {vc:.1f}" if vc >= 1.0 else f"ok, v/c {vc:.2f}")
        d = json.loads((OUT / "data" / "scheme_test.json").read_text())
        drow = next((x for x in d.get("uturn_detour", [])
                     if x["junction"] == code and x["bay"] == side), {})
        one_way = drow.get("one_way_m")
        vals = [f"{side} of junction", conflicting_direction(side), serves,
                one_way if one_way is not None else "beyond drawing",
                drow.get("detour_m") if drow.get("detour_m") is not None else "—",
                round(u["uturn_demand"]), round(u["conflicting_flow"]),
                f"{u['t_c_lo']:.1f}–{u['t_c_hi']:.1f}",
                round(u["cap_optimistic"]), verdict]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(i, j, v); c.font = Font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(i, 10).font = Font(size=9, bold=True,
                                  color=DEFECT if vc >= 1.0 else OKC)
    n = ws.cell(hdr + len(rows) + 2, 1,
                "Ceiling: a single opening passes at most 1,800 veh/h with NO opposing "
                "traffic (3600 / follow-up headway). Demand above that cannot be served "
                "by any bay, metered or not. Gap analysis: HCM form, composition-weighted "
                "critical gap; see scheme_test.json and the dashboard for the full basis.")
    n.font = SUBF
    widths = [14, 12, 40, 13, 13, 12, 14, 12, 12, 16]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def build():
    import json
    bins, _ = parse_all()
    mvall = bins[bins.kind == "movement"]
    days = sorted(mvall.date.unique())
    BOOKS.mkdir(parents=True, exist_ok=True)
    IMGS.mkdir(exist_ok=True)

    # published measurements: bay distances and carriageway widths, never re-derived here
    sch = json.loads((OUT / "data" / "scheme_test.json").read_text())
    det = {}
    for u in sch.get("uturn_detour", []):
        det.setdefault(u["junction"], {})[u["bay"]] = u
    cap = json.loads((OUT / "data" / "capacity.json").read_text())["widths"]

    made, checks = [], []
    for code in sorted(JUNCTIONS, key=lambda c: SCHEME_LABEL[c]):
        arms = JUNCTIONS[code]
        jmv = mvall[mvall.junction == code]
        for di, day in enumerate(days, start=1):
            day_mv = jmv[jmv.date == day]
            derived = di == 2                 # audit finding F: day 2 is derived
            wb = Workbook(); wb.remove(wb.active)
            book_total = 0
            for turn, _tag in GROUPS:         # 4-4-4: L x4, S x4, R x4
                off = {"Left": 1, "Straight": 2, "Right": 3}[turn]
                for fi in range(4):
                    ti = (fi + off) % 4
                    img = IMGS / f"{code}_{turn[0]}{fi}.png"
                    if not img.exists():
                        diagram(arms, fi, ti, turn, img, scheme_route(fi, ti),
                                det.get(code, {}), cap.get(code, {}).get("width_m"))
                    book_total += sheet_for(wb, code, arms, fi, ti, turn,
                                            day_mv, str(day), derived, img,
                                            det.get(code, {}))
            uturn_sheet(wb, code)
            parsed = int(day_mv["count"].sum())
            checks.append((code, di, book_total, parsed))
            name = f"{SCHEME_LABEL[code]}_{code}_Day{di}_{day}.xlsx"
            wb.save(BOOKS / name)
            made.append(name)
    return made, checks


if __name__ == "__main__":
    made, checks = build()
    print("=== Junction turning-movement workbooks ===")
    print("  12 books, mirroring the 12 issued: 6 junctions x 2 days. 12 movement")
    print("  sheets each (4-4-4) plus a U-Turn Bays sheet. Nothing synthetic.\n")
    ok = 0
    for code, di, got, want in checks:
        match = got == want
        ok += match
        print(f"  {SCHEME_LABEL[code]} {code} day {di}  "
              f"vehicles {got:>8,}  parse {want:>8,}  {'OK' if match else 'MISMATCH'}")
    print(f"\n  GATE - books reconciling exactly with the parse: "
          f"**{ok} of {len(checks)}**")
    if ok != len(checks):
        raise SystemExit("a book does not reconcile with the parsed source")
    print(f"\nwritten: {BOOKS}/  ({len(made)} workbooks)")
