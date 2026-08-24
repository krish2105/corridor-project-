"""
The integrity audit and the JSON export.

Both sat near zero coverage — audit.py at 7%, export.py at 0% — and neither is
incidental. audit.py holds the IRC:106 interpolation that IS the project's central
correction, and export.py assembles every number the dashboard shows. They were skipped
because they look IO-shaped, but most of what matters in them is pure.
"""
import json
import pytest


# --- audit.py: the IRC:106 share-dependent factor ---------------------------
# This is the correction the whole engagement turns on. The survey used a static PCU;
# IRC:106 makes it a function of that class's share of the stream. If this interpolation
# is wrong, every corrected figure downstream is wrong with it.

def test_irc_factor_returns_the_low_value_at_or_below_five_percent():
    from src.audit import irc_factor, IRC
    for code, (lo, _hi) in IRC.items():
        assert irc_factor(code, 0.0) == lo
        assert irc_factor(code, 0.05) == lo, "5% is the boundary, inclusive of the low value"
        assert irc_factor(code, 0.01) == lo


def test_irc_factor_returns_the_high_value_at_or_above_ten_percent():
    from src.audit import irc_factor, IRC
    for code, (_lo, hi) in IRC.items():
        assert irc_factor(code, 0.10) == hi, "10% is the boundary, inclusive of the high value"
        assert irc_factor(code, 0.50) == hi
        assert irc_factor(code, 1.0) == hi


def test_irc_factor_interpolates_linearly_between_the_boundaries():
    from src.audit import irc_factor, IRC
    for code, (lo, hi) in IRC.items():
        mid = irc_factor(code, 0.075)
        assert mid == pytest.approx(lo + (hi - lo) / 2), f"{code} is not linear at the midpoint"
        assert lo < mid < hi
        # monotonic across the band
        seq = [irc_factor(code, s / 100) for s in range(5, 11)]
        assert seq == sorted(seq), f"{code} is not monotonic in share"


def test_two_wheelers_at_the_observed_share_require_the_full_irc_uplift():
    """
    The audit's headline: 2W is 49% of this stream and the survey carried it at 0.50,
    where IRC:106 requires 0.75. If this ever stopped returning the high value at the
    observed share, the central finding would quietly evaporate.
    """
    from src.audit import irc_factor
    assert irc_factor("TWO_W", 0.4911) == 0.75


def test_every_correctable_class_has_an_irc_band_and_no_composite_does():
    """
    CLAUDE.md's rule: correct only what maps 1:1, band the rest. A composite appearing in
    IRC would mean inventing a point estimate for a bucket that mixes classes.
    """
    from src.audit import IRC, COMPOSITE
    assert not (set(IRC) & COMPOSITE), "a composite class has been given a point factor"
    for lo, hi in IRC.values():
        assert lo < hi, "the high-share factor must exceed the low-share one"


def test_say_records_every_line_it_prints(capsys):
    from src import audit
    before = len(audit.L)
    audit.say("a line")
    audit.say()
    assert audit.L[before:] == ["a line", ""]
    assert "a line" in capsys.readouterr().out


# --- audit.py: the conservation and balance checks --------------------------
def test_conservation_check_runs_and_reports_on_real_shaped_data(synth_bins):
    """
    Exercises the IN_*/OUT_* vs movements comparison. It was never imported by a test, so
    a change to the groupby keys would have gone unnoticed until a report came out wrong.
    """
    from src import audit
    before = len(audit.L)
    audit.check_conservation(synth_bins)
    out = "\n".join(audit.L[before:])
    assert "Conservation" in out
    assert "movements" in out.lower()


def test_balance_check_runs_and_reports(synth_bins):
    from src import audit
    before = len(audit.L)
    audit.check_balance(synth_bins)
    assert len(audit.L) > before


def test_day2_check_runs_on_two_dated_days(synth_bins):
    from src import audit
    if synth_bins.date.nunique() < 2:
        pytest.skip("fixture carries one survey day")
    before = len(audit.L)
    audit.check_day2(synth_bins)
    assert len(audit.L) > before


# --- export.py: Douglas-Peucker line simplification -------------------------
# The basemap ships 2,161 surveyed features to a phone. Simplification is what makes that
# viable, and getting it wrong moves surveyed geometry on the map without any error.

def test_simplify_collapses_a_straight_line_to_its_endpoints():
    from src.export import _simplify
    line = [(0.0, 0.0), (0.001, 0.0), (0.002, 0.0), (0.003, 0.0)]
    assert _simplify(line) == [(0.0, 0.0), (0.003, 0.0)]


def test_simplify_keeps_a_vertex_that_deviates_more_than_the_tolerance():
    from src.export import _simplify
    # a detour far larger than the ~0.5 m tolerance must survive
    line = [(0.0, 0.0), (0.001, 0.01), (0.002, 0.0)]
    assert _simplify(line) == line, "a real corner was flattened away"


def test_simplify_drops_a_vertex_inside_the_tolerance():
    from src.export import _simplify
    tol = 0.0000045
    line = [(0.0, 0.0), (0.001, tol / 4), (0.002, 0.0)]
    assert _simplify(line) == [(0.0, 0.0), (0.002, 0.0)]


def test_simplify_never_moves_the_endpoints():
    """
    A simplified line that starts or ends somewhere else is a surveyed feature relocated
    on the map, which no reader could detect by looking.
    """
    from src.export import _simplify
    line = [(75.78, 26.91), (75.781, 26.9101), (75.782, 26.9099), (75.783, 26.91)]
    out = _simplify(line)
    assert out[0] == line[0] and out[-1] == line[-1]
    assert len(out) <= len(line)


def test_simplify_passes_through_degenerate_input():
    from src.export import _simplify
    assert _simplify([]) == []
    assert _simplify([(1.0, 2.0)]) == [(1.0, 2.0)]
    assert _simplify([(1.0, 2.0), (3.0, 4.0)]) == [(1.0, 2.0), (3.0, 4.0)]


def test_simplify_survives_a_line_long_enough_to_blow_the_default_recursion():
    """
    It raises the recursion limit and restores it. A drawing line with thousands of
    vertices is ordinary in this dataset.
    """
    import sys
    from src.export import _simplify
    before = sys.getrecursionlimit()
    zig = [(i * 1e-5, (i % 2) * 1e-3) for i in range(3000)]
    out = _simplify(zig)
    assert sys.getrecursionlimit() == before, "the recursion limit was left raised"
    assert out[0] == zig[0] and out[-1] == zig[-1]


# --- export.py: the heavy/summary payload split -----------------------------
def test_split_moves_the_heavy_series_out_and_leaves_a_pointer(tmp_path, monkeypatch):
    """
    corridor.json is what a JDA officer downloads on a phone. The split is what keeps it
    ~88 KB instead of shipping every 15-minute series inline. The summary must keep a
    hint of what it dropped, or the page cannot know to fetch it.
    """
    from src import export
    monkeypatch.setattr(export, "OUT_DATA", tmp_path)
    (tmp_path / "profiles.json").write_text(json.dumps({
        "mean_hours_over": 9, "los_grid": [{"a": 1}], "cumulative": [{"b": 2}]}))
    summary, heavy = export._split("profiles")
    assert set(heavy) == {"los_grid", "cumulative"}
    assert "los_grid" not in summary and "cumulative" not in summary
    assert summary["mean_hours_over"] == 9
    assert summary["series_available"] == ["cumulative", "los_grid"]


def test_split_returns_none_when_the_stage_has_not_run(tmp_path, monkeypatch):
    from src import export
    monkeypatch.setattr(export, "OUT_DATA", tmp_path)
    assert export._split("profiles") == (None, None)


def test_split_tolerates_a_dataset_missing_one_of_its_heavy_keys(tmp_path, monkeypatch):
    from src import export
    monkeypatch.setattr(export, "OUT_DATA", tmp_path)
    (tmp_path / "exhibits.json").write_text(json.dumps({"x": 1, "flow_raster": [1, 2]}))
    summary, heavy = export._split("exhibits")
    assert set(heavy) == {"flow_raster"}
    assert summary["series_available"] == ["flow_raster"]


def test_jsonable_serialises_dates_and_refuses_everything_else():
    from datetime import date, datetime
    from src.export import jsonable
    assert jsonable(date(2026, 5, 11)) == "2026-05-11"
    assert jsonable(datetime(2026, 5, 11, 9, 30)).startswith("2026-05-11T09:30")
    with pytest.raises(TypeError):
        jsonable(object())


def test_optional_stage_loaders_return_none_before_their_stage_runs(tmp_path, monkeypatch):
    from src import export
    monkeypatch.setattr(export, "OUT_DATA", tmp_path)
    assert export._scheme() is None
    assert export._profiles() is None
    assert export._exhibits() is None


# --- audit.py: the two gates that need a real workbook ----------------------
# These are the gates rewritten in Tier 1 — PCU constancy across all 96 intervals, and
# the peak compared against the workbook's own rolling-hour sheets. Both open .xlsx
# directly, which is exactly why neither had ever been exercised.


def test_pcu_interval_check_finds_no_failure_when_factors_really_are_static(synth_workbook):
    """
    The interval test: apply the factors to every class count on every fifteen-minute row
    and compare against that row's own stored Grand Total (PCU's). The fixture is built
    WITH static factors, so a correct gate finds zero failures across every sheet.
    """
    from src.audit import pcu_interval_check
    from src.pcu import SURVEYED
    checked, failed, _d, _w = pcu_interval_check(
        SURVEYED, source=synth_workbook.parent.parent, dirs=[synth_workbook.parent.name])
    assert checked == 288, f"expected 3 sheets x 96 rows, got {checked}"
    assert failed == 0


def test_pcu_interval_check_catches_a_single_perturbed_interval(synth_workbook, tmp_path):
    """
    The gate's whole purpose, and what makes it worth having: it must fail on ONE bad row
    out of 288. A day-total check could not — a factor that varied between intervals
    still sums to a single count-weighted ratio, which is exactly why this replaced it.
    """
    import shutil
    from openpyxl import load_workbook
    from src.audit import pcu_interval_check
    from src.pcu import SURVEYED
    from src.tmc_parse import COL_PCU, ROW_BINS
    # work on a copy: synth_workbook is session-scoped and a sibling test reads it clean
    d = tmp_path / "INT_11-05-2026"
    d.mkdir()
    target = d / synth_workbook.name
    shutil.copy(synth_workbook, target)
    wb = load_workbook(target)
    ws = wb["V_1"]
    r = list(ROW_BINS)[10]
    ws.cell(row=r, column=COL_PCU, value=ws.cell(row=r, column=COL_PCU).value + 5)
    wb.save(target)
    wb.close()

    _c, failed, delta, where = pcu_interval_check(
        SURVEYED, source=tmp_path, dirs=[d.name])
    assert failed == 1, f"a perturbed interval slipped past the gate ({failed} found)"
    assert delta == pytest.approx(5.0)
    assert "V_1" in where and str(r) in where


def test_workbook_rolling_peaks_finds_the_hour_the_workbook_itself_calls_busiest(
        synth_workbook):
    """
    The half of the peak-hour gate that reads THEIR answer rather than re-deriving ours.
    ROW_HOURS was declared in tmc_parse.py and opened by no module, so a section titled
    "re-derived vs the workbook's stated peaks" compared against nothing.

    The fixture puts a 2x spike in one four-bin window, so the maximum is unambiguous.
    """
    from src.audit import workbook_rolling_peaks
    rows = workbook_rolling_peaks(source=synth_workbook.parent.parent,
                                  dirs=[synth_workbook.parent.name])
    assert len(rows) == 1
    r = rows[0]
    assert r["wb_peak_window"] == "W04", (
        f"the spike sits in window W04; the reader picked {r['wb_peak_window']}")
    # four peak bins at 200 fast + 50 slow per class across 10 classes
    assert r["wb_peak_veh"] > 0
    assert r["workbook"].startswith("01_TMC")


def test_workbook_rolling_peaks_skips_a_workbook_with_no_total_in_sheet(tmp_path):
    """A workbook without the approach-total sheet has no rolling hours to read."""
    from openpyxl import Workbook
    from src.audit import workbook_rolling_peaks
    d = tmp_path / "INT_11-05-2026"
    d.mkdir()
    wb = Workbook()
    wb.active.title = "V_1"
    wb.save(d / "99_TMC.xlsx")
    wb.close()
    assert workbook_rolling_peaks(source=tmp_path, dirs=[d.name]) == []


def test_workbook_rolling_peaks_returns_nothing_when_no_workbooks_exist(tmp_path):
    from src.audit import workbook_rolling_peaks
    (tmp_path / "empty").mkdir()
    assert workbook_rolling_peaks(source=tmp_path, dirs=["empty"]) == []
