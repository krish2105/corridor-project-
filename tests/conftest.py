"""
Synthetic survey fixtures.

WHY THESE EXIST
Coverage of the analysis layer was 23-46% and the uncovered part was the part that
matters: `approach_pcu`, `analyse`, `scenarios`, `uturn_verdict`, `oversaturated_hours`,
`through_vs_turning`, `corridor_order` — every function that turns the workbooks into a
published number. Their bodies were never executed by any test. A silent error in
`approach_pcu` would flow into both artifacts, the dashboard, the capacity report and the
commercial pack, and nothing would fail.

They were untested because they need the client's twelve workbooks and a 198 MB CAD
drawing, neither of which is in this repository by design. So the fixture builds a
corridor small enough to reason about by hand: two junctions, one day, four arms each,
twelve movements, ninety-six fifteen-minute bins, with counts chosen so the expected
answers can be worked out on paper and asserted exactly.

WHAT IS DELIBERATELY NOT SYNTHETIC
The class scheme, the arm ordering, the 08:00 day boundary and the bin structure are
copied from the real survey rather than invented, because a fixture that disagrees with
the source format tests the fixture instead of the code.
"""
import json
from datetime import date, datetime, timedelta

import pandas as pd
import pytest

DAY = date(2026, 5, 11)
START = datetime(2026, 5, 11, 8, 0)
N_BINS = 96                      # 24 h of 15-minute bins, 08:00 -> 08:00

# REAL junction codes with their REAL arm names, carrying SYNTHETIC counts.
#
# Not a stylistic choice. analyse.py and scheme_test.py resolve arm order by looking the
# junction code up in the module-level JUNCTIONS / JUNCTION_COORDS registries in
# config.py, so a junction that is not in those dicts raises KeyError before any
# arithmetic runs. Invented codes fail immediately.
#
# That coupling is exactly why this layer had no tests: the analysis functions cannot be
# exercised on data that does not describe this specific corridor. Taking the arms as a
# parameter would fix it properly and is recorded as a finding; using the real codes here
# buys the coverage now without refactoring the analysis right before a client sees it.
from src.config import JUNCTIONS

ARMS = {code: list(JUNCTIONS[code]) for code in ("TMC-01", "TMC-02")}
MOVES = ["Left", "Straight", "Right"]

# A flat stream: every bin identical, so peak-hour arithmetic has an exact answer and any
# drift shows up immediately. Composition is roughly the real one - two-wheelers about
# half the stream - because the PCU correction is share-dependent and a fixture with an
# unrealistic mix would exercise the wrong branch of factor_band().
PER_BIN = {"TWO_W": 50.0, "CAR_BUCKET": 40.0, "AUTO_TRK_BUS": 6.0,
           "AGRI_LCV": 2.0, "CYCLE": 2.0}


def _rows(junction, arms, per_bin, n_bins=N_BINS):
    out = []
    for entry in range(4):
        for offset, mv in enumerate(MOVES, start=1):
            exit_arm = arms[(entry + offset) % 4]
            sheet = f"V_{entry * 3 + offset}"
            for b in range(n_bins):
                t = START + timedelta(minutes=15 * b)
                for cls, n in per_bin.items():
                    out.append(dict(
                        junction=junction, date=DAY, sheet=sheet, kind="movement",
                        arm_from=arms[entry], arm_to=exit_arm, movement=mv,
                        bin_start=pd.Timestamp(t),
                        bin_label=f"{t:%H%M}-{(t + timedelta(minutes=15)):%H%M}",
                        veh_class=cls, count=float(n), stored_pcu=0.0))
    return out


@pytest.fixture(scope="session")
def synth_bins():
    """
    A two-junction corridor with a flat 24-hour profile.

    Exact by construction, per junction:
      each movement    = 96 bins x 100 veh = 9,600 veh/day
      each approach    = 3 movements       = 28,800 veh/day
      junction total   = 4 approaches      = 115,200 veh/day
      any hour         = 4 bins x 100      = 400 veh/movement-hour
    """
    rows = []
    for j, arms in ARMS.items():
        rows += _rows(j, arms, PER_BIN)
    return pd.DataFrame(rows)


@pytest.fixture(scope="session")
def synth_day():
    return DAY


@pytest.fixture(scope="session")
def synth_peaked():
    """
    The same corridor with a real peak, so peak-hour selection has something to find.

    Bins 4-7 (09:00-10:00) carry triple volume on TMC-01 only. Peak hour must therefore
    resolve to 09:00 on TMC-01 and stay flat on TMC-02.
    """
    rows = []
    for j, arms in ARMS.items():
        base = _rows(j, arms, PER_BIN)
        if j == "TMC-01":
            for r in base:
                b = int((r["bin_start"] - pd.Timestamp(START)).total_seconds() // 900)
                if 4 <= b < 8:
                    r["count"] *= 3.0
        rows += base
    return pd.DataFrame(rows)


# --- synthetic CAD ----------------------------------------------------------
# atlas.py, medians.py and capacity.py stream the DXF as group-code pairs rather than
# loading a document, because the real drawing is 198 MB. That makes a fixture cheap:
# a valid DXF for these readers is just the pairs, in the right sections, on layer names
# the LAYER_CAT map recognises.
#
# Geometry is a straight 500 m corridor running north, with a carriageway edge either
# side 7 m from the centre, a median in the middle broken by one 20 m opening, and two
# buildings. Every answer is therefore known: width 14 m kerb to kerb, one opening, and
# an alignment 500 m long.

_UTM_E, _UTM_N = 578000.0, 2976000.0        # Jaipur, EPSG:32643


def _dxf_entity(layer, points, closed=False):
    out = ["0", "LWPOLYLINE", "8", layer, "90", str(len(points)),
           "70", "1" if closed else "0"]
    for x, y in points:
        out += ["10", f"{x:.3f}", "20", f"{y:.3f}"]
    return out


@pytest.fixture(scope="session")
def synth_dxf(tmp_path_factory):
    """A minimal DXF the streaming readers accept, with hand-known geometry."""
    n0, n1 = _UTM_N, _UTM_N + 500.0
    pairs = ["0", "SECTION", "2", "ENTITIES"]
    # corridor centreline
    pairs += _dxf_entity("kml road", [(_UTM_E, n0), (_UTM_E, n1)])
    # carriageway edges, 7 m either side -> 14 m kerb to kerb
    pairs += _dxf_entity("BT ROAD", [(_UTM_E - 7, n0), (_UTM_E - 7, n1)])
    pairs += _dxf_entity("BT ROAD", [(_UTM_E + 7, n0), (_UTM_E + 7, n1)])
    # median in two runs with a 20 m opening between 240 m and 260 m
    pairs += _dxf_entity("DIVIDER", [(_UTM_E, n0), (_UTM_E, n0 + 240)])
    pairs += _dxf_entity("DIVIDER", [(_UTM_E, n0 + 260), (_UTM_E, n1)])
    # two buildings clear of the carriageway
    pairs += _dxf_entity("BUILDING", [(_UTM_E + 20, n0 + 50), (_UTM_E + 30, n0 + 50),
                                      (_UTM_E + 30, n0 + 60), (_UTM_E + 20, n0 + 60)],
                         closed=True)
    pairs += _dxf_entity("BUILDING", [(_UTM_E - 30, n0 + 300), (_UTM_E - 20, n0 + 300),
                                      (_UTM_E - 20, n0 + 310), (_UTM_E - 30, n0 + 310)],
                         closed=True)
    pairs += ["0", "ENDSEC", "0", "EOF"]

    path = tmp_path_factory.mktemp("cad") / "synthetic.dxf"
    path.write_text("\n".join(pairs) + "\n")
    return path


def pytest_sessionfinish(session, exitstatus):
    """
    Record how many tests were collected, so the builders do not have to shell out.

    build_pitch.py and service_docs.py previously ran `uv run pytest --collect-only` at
    BUILD time to learn their own test count. It worked, but a document build that spawns
    a test runner is slow, and it breaks anywhere pytest is not installed - which is any
    environment that only wants to render the deliverables.

    ONLY A GREEN, UNFILTERED RUN MAY WRITE IT.

    It also used to run at COLLECTION, before a single test had executed, so the count
    was recorded even when the suite then went red — a headline figure in the README
    asserting a passing suite that had just failed. It runs at session finish now, and
    only on exit status 0. A filtered run - `pytest tests/test_x.py`, `-k`, `-m`,
    or a single node id - collects a handful of tests, and this hook happily recorded that
    handful as the project's test count. The next document build then published "1 tests"
    or "15 tests" as a headline figure in client-facing deliverables. It happened three
    times, and each time it looked like a stale README rather than a corrupting write. CI
    itself runs a filtered step, so the corruption had a scheduled cause.
    """
    from src.config import OUT_DATA
    if exitstatus != 0:
        return                            # a red suite does not get to publish a count
    cfg = session.config
    filtered = (bool(cfg.getoption("keyword", default=""))
                or bool(cfg.getoption("markexpr", default=""))
                or bool(cfg.getoption("file_or_dir", default=[]))
                or bool(getattr(cfg.option, "last_failed", False))
                or bool(getattr(cfg.option, "failed_first", False)))
    if filtered:
        return
    try:
        OUT_DATA.mkdir(parents=True, exist_ok=True)
        n = session.testscollected
        (OUT_DATA / "testcount.json").write_text(json.dumps({"collected": n}))
    except Exception:
        pass          # recording the count must never fail a test run


def needs_generated_output():
    """
    Module-level skip for tests that render a deliverable.

    Rendering any document requires out/data, which is gitignored because every file in
    it derives from the client's workbooks and CAD. On a clean checkout - which is what
    CI runs - those files do not exist and reports._load() calls SystemExit. Sixteen
    tests failed that way on CI's very first run.
    """
    from src.config import OUT_DATA
    return pytest.mark.skipif(
        not (OUT_DATA / "corridor.json").exists(),
        reason="deliverables are generated from client data, absent on a clean checkout")


@pytest.fixture(scope="session")
def published():
    """
    Loader for generated datasets in out/data, or skip.

    Delivered as a fixture rather than an importable helper on purpose: pandas ships a
    top-level `tests` package, so `from tests.conftest import ...` resolves to
    site-packages instead of this file. Fixtures are injected by pytest and dodge the
    shadowing entirely.

    out/ is gitignored because everything in it derives from the client's workbooks and
    CAD, so on a clean checkout - which is exactly what CI runs - these files do not
    exist. A test that reads one without guarding fails for the wrong reason and turns
    CI red on its first run, which is what happened.
    """
    import json
    from src.config import OUT_DATA, ROOT

    # FALL BACK TO web/public BEFORE SKIPPING.
    #
    # 56 tests - 16% of the suite, and precisely the ones binding the published numbers to
    # each other - skipped on every CI run, because out/ is gitignored. Those are the
    # checks that would have caught the queue-cap breach, the retired capacity constant
    # and the mislabelled grid maximum, and not one of them had ever executed in CI.
    #
    # No synthetic fixture is needed. Eleven of the fourteen out/data files are ALREADY
    # committed at web/public/ - the dashboard needs them at build time, so they are in
    # the repo and public. Reading those is real data, not a stand-in, and it cannot drift
    # from the schema because it IS the schema. test_web_public_matches_out_data below
    # keeps the two in step.
    WEBPUB = ROOT / "web" / "public"

    def _load(name):
        for base in (OUT_DATA, WEBPUB):
            p = base / f"{name}.json"
            if p.exists():
                return json.loads(p.read_text())
        pytest.skip(f"{name}.json is not in out/data or web/public")
    return _load


@pytest.fixture(scope="session")
def synth_workbook(tmp_path_factory):
    """
    A workbook with the JDA sheet geometry and synthetic counts.

    audit.py's two hardest gates - the PCU-constancy test across all 96 intervals and the
    peak-hour comparison against the workbooks' own rolling-hour sheets - both open .xlsx
    files directly. That is why audit.py sat at 7% coverage: neither could run without a
    workbook, and the real ones are client data that never enters this repo.

    Built to the same geometry inspect_tmc verified: rows 8-103 are the 96 fifteen-minute
    bins, 104/105 the day totals, 114-206 the 93 rolling hours, columns B-F fast classes,
    G Total Fast, H-L slow, M Total Slow, N Grand Total, O Grand Total PCU. Counts are
    synthetic; the STRUCTURE is real, which is the part the gates read.
    """
    from openpyxl import Workbook
    from src.tmc_parse import (CLASS_COLS, FAST_COLS, SLOW_COLS, ROW_BINS, ROW_HOURS,
                               ROW_TOTAL_VEH, ROW_TOTAL_PCU, COL_TOTAL_FAST,
                               COL_TOTAL_SLOW, COL_GRAND, COL_PCU)
    from src.pcu import SURVEYED

    d = tmp_path_factory.mktemp("wb")
    path = d / "01_TMC (11-05-2026).xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    # one movement sheet and the approach total, enough for both gates
    # IN_1 is what check_pcu back-solves the factors from; TOTAL_IN is what the
    # peak gate reads its rolling hours from; V_1 is a movement sheet.
    for name in ("V_1", "IN_1", "TOTAL_IN"):
        ws = wb.create_sheet(name)
        for i, r in enumerate(ROW_BINS):
            # a peak in the 09:00 hour so the rolling-hour maximum is unambiguous
            base = 200 if 4 <= i < 8 else 100
            fast = slow = 0.0
            pcu = 0.0
            for col, code in CLASS_COLS.items():
                n = base if col in FAST_COLS else base // 4
                ws.cell(row=r, column=col, value=n)
                pcu += n * SURVEYED[code]
                if col in FAST_COLS:
                    fast += n
                else:
                    slow += n
            ws.cell(row=r, column=COL_TOTAL_FAST, value=fast)
            ws.cell(row=r, column=COL_TOTAL_SLOW, value=slow)
            ws.cell(row=r, column=COL_GRAND, value=fast + slow)
            ws.cell(row=r, column=COL_PCU, value=round(pcu, 6))
        # day totals, derived so nothing is a mismatch by construction
        for col in list(CLASS_COLS) + [COL_TOTAL_FAST, COL_TOTAL_SLOW, COL_GRAND, COL_PCU]:
            ws.cell(row=ROW_TOTAL_VEH, column=col,
                    value=sum(ws.cell(row=r, column=col).value or 0 for r in ROW_BINS))
        for col, code in CLASS_COLS.items():
            v = ws.cell(row=ROW_TOTAL_VEH, column=col).value
            ws.cell(row=ROW_TOTAL_PCU, column=col, value=round(v * SURVEYED[code], 6))
        # the rolling-hour block: 93 windows of four consecutive bins
        bins = list(ROW_BINS)
        for k, r in enumerate(ROW_HOURS):
            if k + 3 >= len(bins):
                break
            tot = sum(ws.cell(row=bins[k + j], column=COL_GRAND).value for j in range(4))
            ws.cell(row=r, column=1, value=f"W{k:02d}")
            ws.cell(row=r, column=COL_GRAND, value=tot)
    wb.save(path)
    wb.close()
    return path
